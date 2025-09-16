"""
File I/O utilities for handling CSV and Excel files
Supports chunked reading and memory-efficient processing
"""

import pandas as pd
import io
from typing import Dict, Any, Union, List
import openpyxl


class FileHandler:
    def __init__(self, chunk_size: int = 20000):
        self.chunk_size = chunk_size
        self.preview_rows = 200  # Maximum rows for preview
    
    def load_file(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """
        Load various data file formats and return dict of sheet_name -> DataFrame
        For single-table formats, uses 'main' as the sheet name
        """
        file_extension = uploaded_file.name.lower().split('.')[-1]
        
        if file_extension == 'csv':
            return self._load_csv(uploaded_file)
        elif file_extension in ['xlsx', 'xls']:
            return self._load_excel(uploaded_file)
        elif file_extension == 'json':
            return self._load_json(uploaded_file)
        elif file_extension == 'tsv':
            return self._load_tsv(uploaded_file)
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")
    
    def _load_csv(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Load CSV file with chunked reading for large files"""
        try:
            # Convert uploaded file to BytesIO for pandas compatibility
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            # Convert bytes to StringIO for CSV reading
            if isinstance(file_content, bytes):
                file_content = file_content.decode('utf-8')
            
            csv_buffer = io.StringIO(file_content)
            file_size = len(file_content)
            
            if file_size > 10 * 1024 * 1024:  # 10MB threshold
                # Use chunked reading
                chunks = []
                csv_buffer.seek(0)
                for chunk in pd.read_csv(csv_buffer, chunksize=self.chunk_size):
                    chunks.append(chunk)
                df = pd.concat(chunks, ignore_index=True)
            else:
                # Read entire file
                csv_buffer.seek(0)
                df = pd.read_csv(csv_buffer)
            
            return {'main': df}
            
        except Exception as e:
            raise ValueError(f"Error reading CSV file: {str(e)}")
    
    def _load_excel(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Load Excel file with all sheets"""
        try:
            # Convert uploaded file to BytesIO for pandas compatibility
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            excel_buffer = io.BytesIO(file_content)
            
            # Read all sheets
            sheets_dict = pd.read_excel(excel_buffer, sheet_name=None, engine='openpyxl')
            
            return sheets_dict
            
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
    
    def _load_json(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Load JSON file and convert to DataFrame"""
        try:
            # Convert uploaded file to string
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            if isinstance(file_content, bytes):
                file_content = file_content.decode('utf-8')
            
            import json
            
            # Parse JSON
            json_data = json.loads(file_content)
            
            # Handle different JSON structures
            if isinstance(json_data, list):
                # Array of objects - direct conversion
                df = pd.DataFrame(json_data)
            elif isinstance(json_data, dict):
                # Check if it's a single record or multiple sheets
                if all(isinstance(v, list) for v in json_data.values()):
                    # Multiple arrays - treat as sheets
                    sheets_dict = {}
                    for key, value in json_data.items():
                        sheets_dict[key] = pd.DataFrame(value)
                    return sheets_dict
                else:
                    # Single record - convert to DataFrame
                    df = pd.DataFrame([json_data])
            else:
                raise ValueError("JSON format not supported - expected array of objects or object with arrays")
            
            return {'main': df}
            
        except Exception as e:
            raise ValueError(f"Error reading JSON file: {str(e)}")
    
    def _load_tsv(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Load TSV (Tab-Separated Values) file"""
        try:
            # Convert uploaded file to StringIO for pandas compatibility
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            # Convert bytes to StringIO for CSV reading
            if isinstance(file_content, bytes):
                file_content = file_content.decode('utf-8')
            
            tsv_buffer = io.StringIO(file_content)
            file_size = len(file_content)
            
            if file_size > 10 * 1024 * 1024:  # 10MB threshold
                # Use chunked reading
                chunks = []
                tsv_buffer.seek(0)
                for chunk in pd.read_csv(tsv_buffer, sep='\t', chunksize=self.chunk_size):
                    chunks.append(chunk)
                df = pd.concat(chunks, ignore_index=True)
            else:
                # Read entire file
                tsv_buffer.seek(0)
                df = pd.read_csv(tsv_buffer, sep='\t')
            
            return {'main': df}
            
        except Exception as e:
            raise ValueError(f"Error reading TSV file: {str(e)}")
    
    def get_file_preview(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """
        Get lightweight preview of file with limited rows for UI display
        """
        file_extension = uploaded_file.name.lower().split('.')[-1]
        
        if file_extension == 'csv':
            return self._preview_csv(uploaded_file)
        elif file_extension in ['xlsx', 'xls']:
            return self._preview_excel(uploaded_file)
        elif file_extension == 'json':
            return self._preview_json(uploaded_file)
        elif file_extension == 'tsv':
            return self._preview_tsv(uploaded_file)
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")
    
    def _preview_csv(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Get preview of CSV file with limited rows"""
        try:
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            if isinstance(file_content, bytes):
                file_content = file_content.decode('utf-8')
            
            csv_buffer = io.StringIO(file_content)
            
            # Read only preview_rows for preview
            df = pd.read_csv(csv_buffer, nrows=self.preview_rows)
            
            return {'main': df}
            
        except Exception as e:
            raise ValueError(f"Error previewing CSV file: {str(e)}")
    
    def _preview_excel(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Get preview of Excel file with limited rows per sheet"""
        try:
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            excel_buffer = io.BytesIO(file_content)
            
            # Use openpyxl for streaming read
            workbook = openpyxl.load_workbook(excel_buffer, read_only=True, data_only=True)
            sheets_dict = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Read limited rows for preview
                rows_data = []
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i >= self.preview_rows:
                        break
                    rows_data.append(row)
                
                if rows_data:
                    # Convert to DataFrame
                    df = pd.DataFrame(rows_data[1:], columns=rows_data[0] if rows_data else None)
                    sheets_dict[sheet_name] = df
            
            workbook.close()
            return sheets_dict
            
        except Exception as e:
            raise ValueError(f"Error previewing Excel file: {str(e)}")
    
    def _preview_json(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Get preview of JSON file with limited rows"""
        try:
            # For JSON, load full structure but limit DataFrame rows
            full_data = self._load_json(uploaded_file)
            
            preview_dict = {}
            for sheet_name, df in full_data.items():
                if len(df) > self.preview_rows:
                    preview_dict[sheet_name] = df.head(self.preview_rows)
                else:
                    preview_dict[sheet_name] = df
            
            return preview_dict
            
        except Exception as e:
            raise ValueError(f"Error previewing JSON file: {str(e)}")
    
    def _preview_tsv(self, uploaded_file) -> Dict[str, pd.DataFrame]:
        """Get preview of TSV file with limited rows"""
        try:
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            if isinstance(file_content, bytes):
                file_content = file_content.decode('utf-8')
            
            tsv_buffer = io.StringIO(file_content)
            
            # Read only preview_rows for preview
            df = pd.read_csv(tsv_buffer, sep='\t', nrows=self.preview_rows)
            
            return {'main': df}
            
        except Exception as e:
            raise ValueError(f"Error previewing TSV file: {str(e)}")
    
    def stream_csv_chunks(self, uploaded_file):
        """
        Generator that yields CSV chunks for memory-efficient processing
        """
        try:
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            
            if isinstance(file_content, bytes):
                file_content = file_content.decode('utf-8')
            
            csv_buffer = io.StringIO(file_content)
            
            # Use pandas chunked reading with proper settings for email processing
            chunk_reader = pd.read_csv(
                csv_buffer, 
                chunksize=self.chunk_size,
                dtype=str,  # Keep everything as strings for email processing
                keep_default_na=False,  # Don't convert to NaN
                na_filter=False  # Don't interpret as missing values
            )
            
            for chunk in chunk_reader:
                yield chunk
                
        except Exception as e:
            raise ValueError(f"Error streaming CSV file: {str(e)}")
    
    def stream_excel_chunks(self, uploaded_file, sheet_name: str):
        """
        Generator that yields Excel sheet chunks for memory-efficient processing
        """
        try:
            uploaded_file.seek(0)
            file_content = uploaded_file.getvalue()
            excel_buffer = io.BytesIO(file_content)
            
            # Use openpyxl for streaming read
            workbook = openpyxl.load_workbook(excel_buffer, read_only=True, data_only=True)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            
            sheet = workbook[sheet_name]
            
            # Get header row
            header_row = next(sheet.iter_rows(values_only=True))
            
            # Process in chunks
            chunk_data = []
            for row in sheet.iter_rows(values_only=True, min_row=2):  # Skip header
                chunk_data.append(row)
                
                if len(chunk_data) >= self.chunk_size:
                    # Yield chunk as DataFrame
                    df = pd.DataFrame(chunk_data, columns=header_row)
                    df = df.astype(str)  # Convert to strings for email processing
                    df = df.replace('None', '')  # Replace None values
                    yield df
                    chunk_data = []
            
            # Yield remaining data
            if chunk_data:
                df = pd.DataFrame(chunk_data, columns=header_row)
                df = df.astype(str)
                df = df.replace('None', '')
                yield df
            
            workbook.close()
            
        except Exception as e:
            raise ValueError(f"Error streaming Excel file: {str(e)}")
    
    def save_to_excel(self, sheets_dict: Dict[str, pd.DataFrame]) -> bytes:
        """Save multiple sheets to Excel format and return bytes"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return output.getvalue()
    
    def save_to_csv(self, df: pd.DataFrame) -> str:
        """Save DataFrame to CSV format and return string"""
        return df.to_csv(index=False)
